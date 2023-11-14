#import libraries
from openpyxl import load_workbook

wb = load_workbook("ebayTemplate.xlsx")
sheetNames = wb.sheetnames

for name in sheetNames:
    print(name)

print("Second Cell: ", sheet['B5'].value)